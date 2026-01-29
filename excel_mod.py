import openpyxl
inv=openpyxl.load_workbook("inventory.xlsx")
prod=inv["Sheet1"]
prod_per_supp={}
val_per_supp={}
prod_u10_inven={}
total_cost=prod.cell(1,5)
total_cost.value="Total Value"

for p_row in range(2,prod.max_row+1) :
    supp_name=prod.cell(p_row,4).value
    inventory=prod.cell(p_row,2).value
    cost=prod.cell(p_row,3).value
    pro_no=prod.cell(p_row,1).value
    inv_price=prod.cell(p_row,5)

    if supp_name in prod_per_supp:
        current_supp=prod_per_supp.get(supp_name)
        prod_per_supp[supp_name]=current_supp+1
    else :
        ###print("Adding new supplier")
        prod_per_supp[supp_name] = 1
    if supp_name in val_per_supp:
        c_valu=val_per_supp.get(supp_name)
        val_per_supp[supp_name]=c_valu+inventory*cost
    else:
        val_per_supp[supp_name]=inventory*cost
    if inventory<10 :
        prod_u10_inven[int(pro_no),supp_name]=int(inventory)
    inv_price.value=inventory*cost
print(prod_per_supp)
print(val_per_supp)
print(prod_u10_inven)
inv.save("inv_val.xlsx")